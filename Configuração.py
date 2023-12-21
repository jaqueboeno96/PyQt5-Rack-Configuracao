# -*- coding: utf-8 -*-
"""
Created on Thu Oct 26 11:01:34 2023

@author: CSP1678
"""

import sys
import openpyxl
import pandas as pd
from datetime import date
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QFont
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QLabel, QLineEdit, QPushButton, QFileDialog, QVBoxLayout, QHBoxLayout, QComboBox, QGraphicsView, QGraphicsScene, QGraphicsRectItem, QTableWidget, QTableWidgetItem

def expandir_cores(cores):
    cores_expandidas = []
    index = 0
    cores_adicionadas = 0
    
    while cores_adicionadas < 210:
        # Replicar cada cor quatro vezes
        for _ in range(4):
            cores_expandidas.append(cores[index])
            cores_adicionadas += 1
            if cores_adicionadas >= 210:
                break
        # A cada quatro cores, adicione uma cor "branca" ou "transparente"
        if cores_adicionadas < 210:
            cores_expandidas.append("#FFFFFF")  # ou qualquer outra cor que represente "espaço"
            cores_adicionadas += 1
        index += 1
        
    return cores_expandidas

class ConfiguracaoEstoque(QMainWindow):
    def __init__(self):
        super().__init__()

        self.initUI()

    def initUI(self):
        self.setWindowTitle('Configuração de Estoque')        

        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)
        
        layout = QVBoxLayout()
        
        title_label = QLabel('Depósito racks ABQ3')
        title_font = QFont("Courier New", 14, QFont.Bold)
        title_label.setFont(title_font)
        layout.addWidget(title_label)
        title_label.setAlignment(Qt.AlignCenter)
            
        self.view = QGraphicsView()
        self.scene = QGraphicsScene()
        self.view.setScene(self.scene)
        self.view.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.view.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        layout.addWidget(self.view)

        self.data_label = QLabel()
        layout.addWidget(self.data_label)
        data_font = QFont("Courier New", 11)  
        self.data_label.setFont(data_font)
        
        rack_layout = QHBoxLayout()

        rack_label = QLabel('Racks:')
        rack_font = QFont("Courier New", 11)
        rack_label.setFont(rack_font)
        rack_layout.addWidget(rack_label)

        self.rack_entry = QLineEdit()
        rack_layout.addWidget(self.rack_entry)

        fim_label = QLabel('até:')
        fim_font = QFont("Courier New", 11)
        fim_label.setFont(fim_font)
        rack_layout.addWidget(fim_label)

        self.fim_entry = QLineEdit()
        rack_layout.addWidget(self.fim_entry)
        
        layout.addLayout(rack_layout)

        cor_layout = QHBoxLayout()

        cor_label = QLabel('Característica do Material:')
        cor_font = QFont("Courier New", 11)
        cor_label.setFont(cor_font)
        cor_layout.addWidget(cor_label)

        self.var_opcao = QComboBox()
        self.var_opcao.setFont(cor_font)
        self.var_opcao.addItem("Escolha")
        self.var_opcao.addItem("ABQ3 - Padrão 2,25")
        self.var_opcao.addItem("ABQ3 - Padrão 2,26")
        self.var_opcao.addItem("ABQ3 - Padrão 2,40")
        self.var_opcao.addItem("ABQ3 - Padrão Estreito")
        self.var_opcao.addItem("ABQ3 - Padrão 2,00")
        self.var_opcao.addItem("ABQ3 - Não Padrão")
        self.var_opcao.addItem("ABQ3 - BQD")
        self.var_opcao.addItem("ABQ3 - BQD UPV")
        self.var_opcao.addItem("Descanso/Retrabalho")
        self.var_opcao.setStyleSheet("QComboBox::item { font: 10pt Arial; }")


        cor_layout.addWidget(self.var_opcao)

        layout.addLayout(cor_layout)

        alterar_cor_button = QPushButton('Configurar')
        alterar_cor_button.clicked.connect(self.alterar_cor)
        alterar_cor_font = QFont("Courier New", 11)
        alterar_cor_button.setFont(alterar_cor_font)
        layout.addWidget(alterar_cor_button)

      
        quit_button = QPushButton('Sair')
        quit_button.clicked.connect(self.close)
        quit_font = QFont("Courier New", 11)
        quit_button.setFont(quit_font)
        layout.addWidget(quit_button)
        
        central_widget.setLayout(layout)

        self.rack_width = 27  # Largura de cada rack
        self.rack_height = 125
        self.rack_spacing = 10  # Espaçamento entre os racks

        self.num_racks = 43

        self.racks_colors = self.carregar_configuracao()
        self.empty_cells_count = self.carregar_celulas_vazias()

        self.atualizar_racks_na_interface()
        self.atualizar_data()

    def atualizar_data(self):
        data_atual = date.today()
        self.data_label.setText(f"Data atual: {data_atual}")

    def alterar_cor(self):
        numero_inicio_rack = int(self.rack_entry.text())
        numero_final_rack = int(self.fim_entry.text())

        if 1 <= numero_inicio_rack <= self.num_racks and 1 <= numero_final_rack <= self.num_racks:
            valor = self.var_opcao.currentText()

            cor = self.determinar_cor(valor)

            for rack_numero in range(numero_inicio_rack - 1, numero_final_rack):
                self.racks_colors[rack_numero] = cor

            self.atualizar_racks_na_interface()
            self.salvar_configuracao()
            self.salvar_cores_quadriplicadas()

        else:
            print("Números de rack inválidos.")

    def determinar_cor(self, valor):
        if valor == "ABQ3 - Padrão 2,25":
            return "#00B0F0" 
        elif valor == "ABQ3 - Padrão 2,26":
            return "#FFFF00" 
        elif valor == "ABQ3 - Padrão 2,40":
            return "#FF0000"  
        elif valor == "ABQ3 - Padrão Estreito":
            return "#7030A0"  
        elif valor == "ABQ3 - Padrão 2,00":
            return "#000000"  
        elif valor == "ABQ3 - Não Padrão":
            return "#00B050"  
        elif valor == "ABQ3 - BQD":
            return "#FFC000"  
        elif valor == "ABQ3 - BQD UPV":
            return "#00FF00"  
        elif valor == "Descanso/Retrabalho":
            return "#FF66FF"  
        else:
            return "#FFFFFF"  

    def atualizar_racks_na_interface(self):
        self.scene.clear()
        
        for rack_numero, (cor, empty_cells) in enumerate(zip(self.racks_colors[:self.num_racks], self.empty_cells_count[:self.num_racks])):
            numero_identificacao = rack_numero + 1
            self.criar_rack(rack_numero, cor, numero_identificacao, empty_cells)
            

    def salvar_configuracao(self):
        try:
            wb = openpyxl.load_workbook('configuracao.xlsx')
            ws = wb.active
    
            for i, cor in enumerate(self.racks_colors):
                ws.cell(row=i + 1, column=1, value=cor)
    
            wb.save('configuracao.xlsx')
        except FileNotFoundError:
            print("Arquivo 'configuracao.xlsx' não encontrado.")

    def salvar_cores_quadriplicadas(self):
        try:
            cores_expandidas = expandir_cores(self.racks_colors)
            wb = openpyxl.Workbook()
            ws = wb.active
    
            for i, cor in enumerate(cores_expandidas):
                ws.cell(row=i + 1, column=1, value=cor)
    
            wb.save('cores_quadriplicadas.xlsx')
        except Exception as e:
            print(f"Erro ao salvar cores quadruplicadas: {e}")

    def carregar_configuracao(self):
        try:
            wb = openpyxl.load_workbook('configuracao.xlsx')
            ws = wb.active
            configuracao = [ws.cell(row=i, column=1).value for i in range(1, self.num_racks + 1)]
            return configuracao
        except FileNotFoundError:
            print("Arquivo 'configuracao.xlsx' não encontrado.")
            return ["white"] * self.num_racks
        
    def carregar_celulas_vazias(self):
        try:
            wb = openpyxl.load_workbook('empty_cells_count.xlsx')
            ws = wb.active
            configuracao = [ws.cell(row=i, column=2).value for i in range(2, self.num_racks + 1)]
            return configuracao
        except FileNotFoundError:
            print("Arquivo 'empty_cells_count.xlsx' não encontrado.")
            return [0] * self.num_racks
    
    def criar_rack(self, rack_numero, cor, numero_identificacao, empty_cells):
        x = rack_numero * (self.rack_width + self.rack_spacing)
        y = self.rack_spacing
    
        rack = QGraphicsRectItem(x, y, self.rack_width, self.rack_height)
        brush = rack.brush()
        brush.setColor(QColor(cor))
        brush.setStyle(Qt.SolidPattern)
        rack.setBrush(brush)
        
        font = QFont()
        font.setPointSize(10)
        font.setBold(True)
    
        self.scene.addItem(rack)
             
        label = self.scene.addText(str(numero_identificacao))
        label.setFont(font)
        label.setPos(x + self.rack_width / 2 - label.boundingRect().width() / 2, y - 25)
        
        empty_cells_count = self.scene.addText(f"{empty_cells}")
        empty_cells_count.setFont(font)
        empty_cells_count.setDefaultTextColor(QColor("red"))  # Define a cor do texto
        empty_cells_count.setPos(x + self.rack_width / 2 - empty_cells_count.boundingRect().width() / 2, y + self.rack_height + 5)
        
        if rack_numero == self.num_racks // 2:
            
            font = QFont()
            font.setPointSize(12)
            font.setBold(True)
            
            rack_label = self.scene.addText("Racks")
            rack_label.setFont(font)
            rack_label.setPos(x + self.rack_width / 2 - rack_label.boundingRect().width() / 2, y - 50)
            
            empty_cells_label = self.scene.addText("Vazios")
            empty_cells_label.setFont(font)
            empty_cells_label.setPos(x + self.rack_width / 2 - empty_cells_label.boundingRect().width() / 2, y + self.rack_height + 30)

    def voltar_para_configuracao(self):
        self.hide()
        configuracao_window.show()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    configuracao_window = ConfiguracaoEstoque()

    # Defina uma folha de estilo global para aplicar cores a todos os elementos, exceto o header
    style = """
        QWidget {
            background-color: #ffffff;
            color: #000;
        }
        QLabel, QLineEdit, QPushButton, QComboBox {
            background-color: #c8d0d0;
            color: #000;
        }
    """
    app.setStyleSheet(style)

    configuracao_window.show()
    sys.exit(app.exec_())
